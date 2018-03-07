import urllib.request
import json
import calendar

import config

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment
from openpyxl.utils import (get_column_letter)


# Splits minutes into intervals of 15 minutes
def minute_to_interval(minute):
    if minute < 15:
        return 0
    elif 15 <= minute < 30:
        return 1
    elif 30 <= minute < 45:
        return 2
    else:
        return 3


# Increases interval counter in between start time and end time
def counter_add(interval_counter, start_hour, start_min, end_hour, end_min):
    start_interval = minute_to_interval(start_min)
    end_interval = minute_to_interval(end_min)
    start_idx = (start_hour * 4) + start_interval
    end_idx = (end_hour * 4) + end_interval

    # Passing over midnight
    if end_idx < start_idx:
        for idx, interval in enumerate(interval_counter):
            if start_idx <= idx or end_idx >= idx:
                interval_counter[idx] += 1
    # Normal
    else:
        for idx, interval in enumerate(interval_counter):
            if start_idx <= idx <= end_idx:
                interval_counter[idx] += 1


# Writes the interval counter to the worksheet
def counter_write(worksheet, interval_counter):
    col = 2
    for count in interval_counter:
        worksheet.cell(row=9, column=col).value = count
        col += 1


# Transforms seconds to time string (HH:MM:SS)
def seconds_to_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d:%02d:%02d" % (h, m, s)


# Initializes and formats a new month in the worksheet (merging cells, setting values)
def init_month(worksheet, year, month, current_month_row):
    month_range = calendar.monthrange(year, month)
    merge_row = current_month_row+3
    merge_col = 2
    day = 1

    worksheet.merge_cells(start_row=current_month_row, start_column=2, end_row=current_month_row, end_column=10)
    worksheet.cell(row=current_month_row, column=2).value = calendar.month_name[month]
    worksheet.cell(row=current_month_row, column=2).font = Font(bold=True)

    worksheet.merge_cells(start_row=current_month_row+1, start_column=2, end_row=current_month_row+1, end_column=10)
    worksheet.cell(row=current_month_row+1, column=2).value = "Hours online"
    worksheet.merge_cells(start_row=current_month_row+1, start_column=11, end_row=current_month_row+1, end_column=22)
    worksheet.cell(row=current_month_row+1, column=11).font = Font(bold=True)

    for i in range(0, month_range[1]):
        worksheet.merge_cells(start_row=merge_row, start_column=merge_col, end_row=merge_row, end_column=merge_col+1)
        worksheet.merge_cells(start_row=merge_row+1, start_column=merge_col, end_row=merge_row+1, end_column=merge_col+1)
        worksheet.cell(row=merge_row, column=merge_col).value = day
        worksheet.cell(row=merge_row, column=merge_col).alignment = Alignment(horizontal="center")
        worksheet.cell(row=merge_row + 1, column=merge_col).alignment = Alignment(horizontal="center")
        merge_col += 2
        day += 1

    range_string = "{0}{1}:{2}{3}".format(get_column_letter(2), str(current_month_row+4), get_column_letter(merge_col-2), str(current_month_row+4))
    worksheet.conditional_formatting.add(range_string, ColorScaleRule(start_type='min', start_value=None, start_color='FFFFFF',
                                                 end_type='max', end_value=None, end_color='1F497D'))


# Writes daily values to worksheet
def daily_to_xls(worksheet, day, current_month_row):
    column = day * 2
    worksheet.cell(row=current_month_row+4, column=column).value = 1


# Writes total hours per month to worksheet
def set_month_time(worksheet, current_month_row, current_month_time):
    worksheet.cell(row=current_month_row+1, column=11).value = str(seconds_to_time(current_month_time))
    worksheet.cell(row=current_month_row+1, column=11).font = Font(bold=True)


# Adds additional information / formatting
def add_info(worksheet, name, time, start, obs):
    worksheet.cell(row=2, column=9).value = name
    worksheet.cell(row=3, column=9).value = seconds_to_time(time)
    worksheet.cell(row=2, column=33).value = start
    worksheet.cell(row=3, column=33).value = obs

    worksheet.conditional_formatting.add('B9:CS9', ColorScaleRule(start_type='min', start_value=None, start_color='63BE7B',
                                                            mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                                            end_type='max', end_value=None, end_color='F8696B'))


# Processes the session data
def process_sessions(worksheet, data, name, start_date):
    current_month = 0
    current_month_row = 5
    current_month_time = 0

    total_time = 0
    i = 0
    intervals = 24 * 4
    interval_counter = [0 for x in range(intervals)]
    for session in data:
        time_start = datetime.strptime(session['date'], '%Y-%m-%d %H:%M:%S')
        time_end = datetime.strptime(session['lastupd'], '%Y-%m-%d %H:%M:%S')

        if time_start.month != current_month:
            if current_month != 0:
                set_month_time(worksheet, current_month_row, current_month_time)
            current_month_row += 7
            current_month = time_start.month
            current_year = time_start.year
            init_month(worksheet, current_year, current_month, current_month_row)
            current_month_time = 0
        daily_to_xls(worksheet, time_start.day, current_month_row)
        current_month_time += int(session['time'])

        total_time += int(session['time'])

        counter_add(interval_counter, time_start.hour, time_start.minute, time_end.hour, time_end.minute)

    set_month_time(worksheet, current_month_row, current_month_time)
    counter_write(worksheet, interval_counter)
    add_info(worksheet, name, total_time, start_date, len(data))


# Generates activity report for a given user from a certain point in time (start_date)
def generate_report(template, accountid, name, start_date):
    wb = load_workbook(filename = template)
    worksheet = wb.get_sheet_by_name("Template")
    worksheet.title = name

    # Fetch number of bans for user during given period (from start_date)
    url = "http://meuris.se/cp/getbandata.php?key=%s&acid=%s&date=%s" % (config.key, accountid, start_date)
    response = urllib.request.urlopen(url)
    data = json.loads(response.read().decode('UTF-8'))
    bans = data[0]['count']
    worksheet.cell(row=4, column=9).value = int(bans)

    # Fetch total number of bans for user
    url = "http://meuris.se/cp/getbandata.php?key=%s&acid=%s" % (config.key, accountid)
    response = urllib.request.urlopen(url)
    data = json.loads(response.read().decode('UTF-8'))
    total_bans = data[0]['count']
    worksheet.cell(row=5, column=9).value = int(total_bans)

    # Fetch activity data to generate activity report
    url = "http://meuris.se/cp/getjoindata.php?key=%s&acid=%s&date=%s" % (config.key, accountid, start_date)
    response = urllib.request.urlopen(url)
    data = json.loads(response.read().decode('UTF-8'))
    process_sessions(worksheet, data, name, start_date)

    wb.save("%s\\ActivityReport.xlsx" % config.path)


template = "%s\\ActivityReportTemplate.xlsx" % config.path
generate_report(template, "1", "Jonas", "2017-09-01")